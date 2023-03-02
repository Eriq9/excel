import openpyxl
import pandas as pd
from CoolProp.CoolProp import PropsSI

wrkbk = openpyxl.load_workbook("dane.xlsx")

sh = wrkbk.active
lista = []
listaTemp = []
listaTempPom = []
listaList = []

listaListP = []
listaListQ = []

listaP = []
listaQ = []

listaPPom = []
listaQPom = []

ExcelT = []
ExcelP = []
ExcelQ = []
Excel = []
ExcelPropsi = []
ExcelDeltaT = []
ExcelAlfa = []

x=0
a=77

ExcelT.append("Avg_T")
ExcelP.append("Avg_P")
ExcelQ.append("Avg_Q")
ExcelPropsi.append("PropSI")
ExcelDeltaT.append("DeltaT")
ExcelAlfa.append("Alfa")

def Average(lst):
    return sum(lst) / len(lst)



for i in range(2, sh.max_row + 1):
    cell_obj = sh.cell(row=i, column=6)
    temp = sh.cell(row=i, column=2)
    p = sh.cell(row=i, column=3)
    q = sh.cell(row=i, column=1)
    lista.append(cell_obj.value)

    listaTemp.append(float(temp.value))
    listaTempPom.append(float(temp.value))
    listaP.append(float(p.value))
    listaPPom.append(float(p.value))
    listaQ.append(float(q.value))
    listaQPom.append(float(q.value))

    if lista[i-2] < lista[i-3]:

        listaTemp.pop()
        listaList.append(listaTemp)

        listaP.pop()
        listaListP.append(listaP)

        listaQ.pop()
        listaListQ.append(listaQ)

        print("Temperatury:",listaList)
        print("P:", listaListP)
        print("Q:", listaListQ)

        listaTemp = []
        listaTemp.append(listaTempPom[len(listaTempPom)-1])

        listaP = []
        listaP.append(listaPPom[len(listaPPom) - 1])

        listaQ = []
        listaQ.append(listaQPom[len(listaQPom) - 1])

        x+=1
        print("Średnia T: ",Average(listaList[x-1]))

        print("Średnia P: ", Average(listaListP[x - 1]))

        print("Średnia Q: ", Average(listaListQ[x - 1]))
        print("\n")

        #print(PropsSI('T','P',Average(listaListP[x - 1])*1000,'Q',1,'Water'))
        print("\n")

        ExcelT.append(Average(listaList[x-1]))
        ExcelP.append(Average(listaListP[x - 1]))
        ExcelQ.append(Average(listaListQ[x - 1]))
        ExcelPropsi.append(PropsSI('T','P',Average(listaListP[x - 1])*1000,'Q',1,'Water'))
        ExcelDeltaT.append(Average(listaList[x-1])-PropsSI('T','P',Average(listaListP[x - 1])*1000,'Q',1,'Water'))
        ExcelAlfa.append(Average(listaListQ[x - 1])/(a*(Average(listaList[x-1])-PropsSI('T','P',Average(listaListP[x - 1])*1000,'Q',1,'Water'))))

def SaveToExcel():


    Excel.append(ExcelT)
    Excel.append(ExcelP)
    Excel.append(ExcelQ)
    Excel.append(ExcelPropsi)
    Excel.append(ExcelDeltaT)
    Excel.append(ExcelAlfa)

    my_data = pd.DataFrame(Excel)
    my_data.transpose().to_excel("Avg_File6.xlsx",index=False,header=False)

SaveToExcel()


